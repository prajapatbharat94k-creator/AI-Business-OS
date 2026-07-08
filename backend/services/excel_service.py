import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

def generate_report_excel(report_type: str, data: list) -> bytes:
    """
    Generates an Excel workbook for the specified report type and returns the file as bytes.
    report_type: 'sales' | 'inventory' | 'customer' | 'employee'
    data: list of dicts with the report rows
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{report_type.capitalize()} Report"
    
    # Enable grid lines explicitly
    ws.views.sheetView[0].showGridLines = True
    
    # Styled headers definition
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # slate-800
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    border_thin = Side(border_style="thin", color="CBD5E1") # slate-300
    border_double = Side(border_style="double", color="1E293B")
    cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    
    if report_type == "sales":
        headers = ["Date", "Invoice #", "Customer", "Branch", "Payment Method", "Status", "Subtotal", "GST", "Total"]
        ws.append(headers)
        
        for row in data:
            ws.append([
                row.get("date", ""),
                row.get("invoice_number", ""),
                row.get("customer_name", ""),
                row.get("branch_name", ""),
                row.get("payment_method", ""),
                row.get("status", ""),
                row.get("subtotal", 0.0),
                row.get("gst_amount", 0.0),
                row.get("total_amount", 0.0)
            ])
            
        # Apply formatting to currency columns (G, H, I) and date (A)
        for row_idx in range(2, len(data) + 2):
            ws.cell(row=row_idx, column=1).number_format = 'yyyy-mm-dd hh:mm'
            ws.cell(row=row_idx, column=7).number_format = '$#,##0.00'
            ws.cell(row=row_idx, column=8).number_format = '$#,##0.00'
            ws.cell(row=row_idx, column=9).number_format = '$#,##0.00'
            
    elif report_type == "inventory":
        headers = ["Product Name", "SKU / Barcode", "Price", "Cost Price", "Stock Qty", "Min Threshold", "Status", "Category", "Branch", "Supplier"]
        ws.append(headers)
        
        for row in data:
            qty = row.get("stock_quantity", 0)
            thresh = row.get("low_stock_threshold", 10)
            status = "Low Stock" if qty <= thresh else "In Stock"
            
            ws.append([
                row.get("name", ""),
                row.get("sku", ""),
                row.get("price", 0.0),
                row.get("cost_price", 0.0),
                qty,
                thresh,
                status,
                row.get("category_name", ""),
                row.get("branch_name", ""),
                row.get("supplier_name", "")
            ])
            
        # Apply currency formatting (C, D)
        for row_idx in range(2, len(data) + 2):
            ws.cell(row=row_idx, column=3).number_format = '$#,##0.00'
            ws.cell(row=row_idx, column=4).number_format = '$#,##0.00'
            
    elif report_type == "customer":
        headers = ["Customer Name", "Email", "Phone", "Address", "Loyalty Points", "Total Orders", "Total Spent"]
        ws.append(headers)
        
        for row in data:
            ws.append([
                row.get("name", ""),
                row.get("email", ""),
                row.get("phone", ""),
                row.get("address", ""),
                row.get("loyalty_points", 0),
                row.get("total_orders", 0),
                row.get("total_spent", 0.0)
            ])
            
        # Format spent (G)
        for row_idx in range(2, len(data) + 2):
            ws.cell(row=row_idx, column=7).number_format = '$#,##0.00'
            
    elif report_type == "employee":
        headers = ["Employee Name", "Email", "Position", "Branch", "Salary", "Hire Date", "Leave Balance"]
        ws.append(headers)
        
        for row in data:
            ws.append([
                row.get("name", ""),
                row.get("email", ""),
                row.get("position", ""),
                row.get("branch_name", ""),
                row.get("salary", 0.0),
                row.get("hire_date", ""),
                row.get("leave_balance", 20)
            ])
            
        # Format Salary (E) and date (F)
        for row_idx in range(2, len(data) + 2):
            ws.cell(row=row_idx, column=5).number_format = '$#,##0.00'
            ws.cell(row=row_idx, column=6).number_format = 'yyyy-mm-dd'
            
    else:
        # Default empty headers
        ws.append(["No Data"])
        
    # Styling columns and headers
    for col in ws.columns:
        # Header style
        cell = col[0]
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=Side(style="medium", color="1E293B"))
        
        # Determine column width based on max content length
        max_len = len(str(cell.value or ''))
        for cell_in_col in col[1:]:
            cell_in_col.font = Font(name="Calibri", size=10)
            cell_in_col.border = cell_border
            cell_in_col.alignment = align_left
            val_str = str(cell_in_col.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
                
            # Align numeric cells to right
            if isinstance(cell_in_col.value, (int, float)):
                cell_in_col.alignment = align_right
                
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 3, 12)
        
    # Write workbook to bytes stream
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return file_stream.getvalue()
